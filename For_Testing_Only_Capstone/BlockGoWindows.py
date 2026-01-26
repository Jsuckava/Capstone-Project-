# -*- coding: utf-8 -*-
"""
SECURE STUDENT FEE SYSTEM – MS SQL SERVER (WINDOWS)
• Uses pyodbc + SQL Server
• Polls Notifications table (no LISTEN/NOTIFY)
• Full tamper detection, encryption, GUI
"""
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
import pyodbc
import hashlib
import json
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import threading
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad, unpad
from Crypto.Random import get_random_bytes
import socket

DB_CONFIG = {
    'driver': '{ODBC Driver 17 for SQL Server}',
    'server': '',       
    'database': '',  
    'trusted_connection': 'yes'    
  
}

AES_KEY = b'ThisIsA32ByteLongEncryptionKeys!'

def encrypt_data(plaintext: str) -> str:
    iv = get_random_bytes(16)
    cipher = AES.new(AES_KEY, AES.MODE_CBC, iv)
    ct = cipher.encrypt(pad(plaintext.encode(), AES.block_size))
    return (iv + ct).hex()

def decrypt_data(ct_hex: str) -> str:
    try:
        raw = bytes.fromhex(ct_hex)
        iv, ct = raw[:16], raw[16:]
        cipher = AES.new(AES_KEY, AES.MODE_CBC, iv)
        return unpad(cipher.decrypt(ct), AES.block_size).decode()
    except Exception:
        return "[Decrypt error]"

def hash_sha256(text: str) -> str:
    return hashlib.sha256(text.encode()).hexdigest()

class Database:
    def __init__(self):
        conn_str = (
            f"DRIVER={DB_CONFIG['driver']};"
            f"SERVER={DB_CONFIG['server']};"
            f"DATABASE={DB_CONFIG['database']};"
            f"Trusted_Connection={DB_CONFIG.get('trusted_connection', 'no')};"
        )
        if 'uid' in DB_CONFIG:
            conn_str += f"UID={DB_CONFIG['uid']};PWD={DB_CONFIG['pwd']};"
        self.conn = pyodbc.connect(conn_str, autocommit=True)
        self.cur = self.conn.cursor()
        self.last_seen_id = self._get_max_id()
        self.last_notification_id = self._get_max_notification_id()

    def _get_max_id(self):
        self.cur.execute("SELECT MAX(transactionId) FROM StudentData")
        row = self.cur.fetchone()
        return row[0] if row and row[0] else 0

    def _get_max_notification_id(self):
        self.cur.execute("SELECT MAX(id) FROM Notifications")
        row = self.cur.fetchone()
        return row[0] if row and row[0] else 0

    def verify_user(self, username, password):
        self.cur.execute("SELECT userName, Password, role FROM Users WHERE userName = ?", username)
        row = self.cur.fetchone()
        if row and row.Password == hash_sha256(password):
            return True, row.role
        return False, None

    def insert_transaction(self, name, section, course, year, address,
                           amount, transaction_type, created_by):
        enc = [encrypt_data(x) for x in (name, section, course, year, address)]
        self.cur.execute("SELECT TOP 1 hash_curr FROM StudentData ORDER BY transactionId DESC")
        row = self.cur.fetchone()
        hash_prev = row.hash_curr if row else '0'*64
        concat = "".join(enc) + str(amount) + transaction_type + created_by + hash_prev
        hash_curr = hash_sha256(concat)

        sql = """
            INSERT INTO StudentData(name, section, course, year, address, Amount,
                                    TransactionCreated, created_by, hash_prev, hash_curr)
            VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        self.cur.execute(sql, *enc, amount, transaction_type, created_by, hash_prev, hash_curr)
        
        self.cur.execute("SELECT @@IDENTITY")
        new_id = int(self.cur.fetchone()[0])
        
        msg = f"DATA CREATED by {created_by} on transactionID {new_id}"
        self._notify(msg)
        return new_id

    def update_transaction(self, tid, column, new_val, username):
        self.cur.execute("SELECT * FROM StudentData WHERE transactionId = ?", tid)
        row = self.cur.fetchone()
        if not row: return False

        cols = ['name','section','course','year','address','TransactionCreated']
        enc_vals = [getattr(row, c) for c in cols]
        if column in cols:
            enc_vals[cols.index(column)] = encrypt_data(new_val)

        hash_prev = row.hash_prev
        amount_val = new_val if column == 'Amount' else row.Amount
        concat = "".join(enc_vals) + str(amount_val) + row.created_by + hash_prev
        hash_curr = hash_sha256(concat)
        old_hash = row.hash_curr

        sql = f"UPDATE StudentData SET {column} = ?, hash_curr = ? WHERE transactionId = ?"
        val = encrypt_data(new_val) if column in cols else new_val
        self.cur.execute(sql, val, hash_curr, tid)

        if old_hash != hash_curr:
            msg = f"DATA MODIFIED! – transaction {tid} updated by {username}"
            self._notify(msg)
        return True

    def delete_transaction(self, tid, username):
        self.cur.execute("DELETE FROM StudentData WHERE transactionId = ?", tid)
        self._notify(f"DATA DELETED – transaction {tid} by {username}")

    def get_all(self):
        self.cur.execute("SELECT * FROM StudentData ORDER BY transactionId")
        return [self._row_to_dict(r) for r in self.cur.fetchall()]

    def get_all_raw(self):
        self.cur.execute("SELECT * FROM StudentData ORDER BY transactionId")
        rows = self.cur.fetchall()
        cols = [column[0] for column in self.cur.description]
        return [dict(zip(cols, row)) for row in rows]

    def search_by_name(self, plain_name):
        enc_name = encrypt_data(plain_name)
        self.cur.execute("SELECT * FROM StudentData WHERE name = ? ORDER BY transactionId", enc_name)
        return [self._row_to_dict(r) for r in self.cur.fetchall()]

    def _row_to_dict(self, row):
        return {
            'transactionid': row.transactionId,
            'name': decrypt_data(row.name),
            'section': decrypt_data(row.section),
            'course': decrypt_data(row.course),
            'year': decrypt_data(row.year),
            'address': decrypt_data(row.address),
            'amount': float(row.Amount) if row.Amount else 0,
            'TransactionCreated': row.TransactionCreated,
            'created_by': row.created_by,
            'hash_prev': row.hash_prev,
            'hash_curr': row.hash_curr
        }

    def _notify(self, msg):
        self.cur.execute("INSERT INTO Notifications (message) VALUES (?)", msg)

    def get_notifications(self):
        self.cur.execute("SELECT id, message FROM Notifications WHERE id > ? ORDER BY id", self.last_notification_id)
        rows = self.cur.fetchall()
        notifs = []
        for r in rows:
            notifs.append({"type": "activity", "msg": r.message})
            if r.id > self.last_notification_id:
                self.last_notification_id = r.id
        return notifs

    def get_daily_user_report(self, target_date=None):
        if target_date is None:
            target_date = datetime.now().date()
        else:
            target_date = datetime.strptime(target_date, "%Y-%m-%d").date()

        sql = """
            SELECT created_by, name, section, course, year, address, Amount, TransactionCreated, transactionId
            FROM StudentData
            WHERE CAST(transactionId AS VARCHAR) LIKE ? + '%'
              AND created_by IN (SELECT userName FROM Users WHERE role != 'admin')
            ORDER BY created_by, transactionId
        """
        self.cur.execute(sql, str(target_date))
        rows = self.cur.fetchall()
        report = {}
        for r in rows:
            user = r.created_by
            if user not in report:
                report[user] = []
            report[user].append({
                'ID': r.transactionId,
                'Name': decrypt_data(r.name),
                'Section': decrypt_data(r.section),
                'Course': decrypt_data(r.course),
                'Year': decrypt_data(r.year),
                'Address': decrypt_data(r.address),
                'Amount': float(r.Amount),
                'Type': r.TransactionCreated
            })
        return report, target_date

class App:
    def __init__(self, master):
        self.master = master
        self.master.title("Secure Student System - MS SQL")
        self.master.geometry("1350x720")
        self.db = Database()
        self.role = None
        self.user = None
        self.editable_row = None
        self.create_login_ui()

        threading.Thread(target=self.tamper_check, daemon=True).start()
        self.master.after(1000, self.check_notifications)  # Poll every 1s

    def clear_ui(self):
        for w in self.master.winfo_children():
            w.destroy()

    def create_login_ui(self):
        self.clear_ui()
        f = ttk.Frame(self.master, padding=30); f.pack(expand=True)
        ttk.Label(f, text="Username").grid(row=0, column=0, sticky='w', pady=6)
        self.e_user = ttk.Entry(f, width=30); self.e_user.grid(row=0, column=1, pady=6)
        ttk.Label(f, text="Password").grid(row=1, column=0, sticky='w', pady=6)
        self.e_pass = ttk.Entry(f, width=30, show='*'); self.e_pass.grid(row=1, column=1, pady=6)
        ttk.Button(f, text="Login", command=self.login).grid(row=2, column=0, columnspan=2, pady=15)

    def login(self):
        u, p = self.e_user.get().strip(), self.e_pass.get()
        ok, role = self.db.verify_user(u, p)
        if ok:
            self.user, self.role = u, role
            desktop = socket.gethostname()
            self.db.cur.execute(
                "INSERT INTO LoginData (DesktopName) VALUES (?)",
                desktop
            )
            self.create_dashboard()
        else:
            messagebox.showerror("Login", "Invalid credentials")

    def create_dashboard(self):
        self.clear_ui()
        top = ttk.Frame(self.master); top.pack(fill='x', padx=12, pady=6)
        ttk.Label(top, text="Search student name:").pack(side='left')
        self.search_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.search_var, width=30).pack(side='left', padx=4)
        ttk.Button(top, text="Search", command=self.do_search).pack(side='left', padx=2)
        ttk.Button(top, text="Clear", command=self.clear_search).pack(side='left', padx=2)
        ttk.Button(top, text="Add Transaction", command=self.add_dialog).pack(side='right', padx=5)
        ttk.Button(top, text="Change Password", command=self.change_password_dialog).pack(side='right', padx=5)
        if self.role == 'admin':
            ttk.Button(top, text="Add User", command=self.add_user_dialog).pack(side='right', padx=5)
            ttk.Button(top, text="Export Daily Report", command=self.export_daily_report).pack(side='right', padx=5)
        ttk.Button(top, text="Logout", command=self.logout).pack(side='right', padx=5)

        style = ttk.Style(self.master)
        style.theme_use("default")
        style.configure("Treeview", background="white", foreground="black", rowheight=25, fieldbackground="white")
        style.map('Treeview', background=[('selected', 'lightblue')])

        cols = ("ID","Name","Section","Course","Year","Address","Amount","Transaction Type","Created By")
        self.tree = ttk.Treeview(self.master, columns=cols, show='headings')
        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=110, anchor='center')
        self.tree.column("Name", width=160)
        self.tree.column("Address", width=180)
        self.tree.column("Transaction Type", width=150)
        self.tree.pack(fill='both', expand=True, padx=12, pady=6)
        self.tree.bind("<Double-1>", self.on_double_click)

        if self.role == 'admin':
            admin_bar = ttk.Frame(self.master); admin_bar.pack(fill='x', pady=2)
            ttk.Button(admin_bar, text="Delete Selected", command=self.delete_selected).pack(side='right', padx=5)

        self.load_all()

    def load_all(self):
        data = self.db.get_all()
        self._populate_tree(data)

    def do_search(self):
        term = self.search_var.get().strip()
        if not term:
            messagebox.showinfo("Search", "Enter a name first.")
            return
        data = self.db.search_by_name(term)
        self._populate_tree(data)

    def clear_search(self):
        self.search_var.set("")
        self.load_all()

    def _populate_tree(self, data):
        self.tree.delete(*self.tree.get_children())
        for r in data:
            self.tree.insert('', 'end', values=(
                r['transactionid'], r['name'], r['section'], r['course'],
                r['year'], r['address'], r['amount'], r['TransactionCreated'], r['created_by']
            ))

    def add_dialog(self):
        dlg = AddTransactionDialog(self.master, self.user)
        self.master.wait_window(dlg)
        if dlg.result:
            self.load_all()

    def on_double_click(self, event):
        sel = self.tree.selection()
        if not sel: return
        item = self.tree.item(sel[0])
        col = self.tree.identify_column(event.x)
        col_idx = int(col[1:]) - 1
        col_name = self.tree['columns'][col_idx]
        tid = item['values'][0]

        if self.role == 'user':
            if self.editable_row and self.editable_row != tid:
                messagebox.showwarning(" gEdit", "Finish the current edit first.")
                return
            self.editable_row = tid

        if self.role == 'admin':
            self.start_cell_edit(sel[0], col_idx, col_name, tid)
            return

        new_val = simpledialog.askstring("Edit", f"New value for {col_name} (ID {tid}):")
        if new_val is None: return

        col_map = {
            'Name': 'name', 'Section': 'section', 'Course': 'course',
            'Year': 'year', 'Address': 'address', 'Amount': 'Amount',
            'Transaction Type': 'TransactionCreated'
        }
        db_col = col_map.get(col_name)
        if not db_col: return

        if col_name == 'Amount' and self.role == 'user':
            messagebox.showwarning("Permission", "Users cannot edit Amount.")
            return

        ok = self.db.update_transaction(tid, db_col, new_val, self.user)
        if ok:
            self.load_all()

    def start_cell_edit(self, item_id, col_idx, col_name, tid):
        x, y, w, h = self.tree.bbox(item_id, column=f"#{col_idx + 1}")
        if not x: return
        entry = ttk.Entry(self.tree)
        entry.insert(0, self.tree.item(item_id)['values'][col_idx])
        entry.select_range(0, 'end')
        entry.focus()
        entry.place(x=x, y=y, width=w, height=h)

        col_map = {
            'Name': 'name', 'Section': 'section', 'Course': 'course',
            'Year': 'year', 'Address': 'address', 'Amount': 'Amount',
            'Transaction Type': 'TransactionCreated'
        }

        def save():
            new_val = entry.get()
            db_col = col_map[col_name]
            self.db.update_transaction(tid, db_col, new_val, self.user)
            self.load_all()
            entry.destroy()

        def cancel():
            entry.destroy()

        entry.bind("<Return>", lambda e: save())
        entry.bind("<Escape>", cancel)
        entry.bind("<FocusOut>", lambda e: save())

    def delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Choose a row")
            return
        tid = self.tree.item(sel[0])['values'][0]
        if messagebox.askyesno("Delete", f"Delete ID {tid}?"):
            self.db.delete_transaction(tid, self.user)
            self.load_all()

    def check_notifications(self):
        notes = self.db.get_notifications()
        for n in notes:
            if n.get('type') == 'activity':
                self.show_popup(n['msg'])
        self.master.after(1000, self.check_notifications)

    def show_popup(self, msg):
        p = tk.Toplevel(self.master)
        p.title("System Alert")
        p.geometry("500x140")
        p.configure(bg='#ffcccc' if 'TAMPERED' in msg else '#ccffcc' if 'CREATED' in msg else '#fff9cc')
        ttk.Label(p, text=msg, wraplength=480, justify='center', font=('Arial', 10, 'bold')).pack(pady=20)
        ttk.Button(p, text="OK", command=p.destroy).pack(pady=5)
        p.after(8000, p.destroy)

    def logout(self):
        self.user = self.role = None
        self.create_login_ui()

    def add_user_dialog(self):
        dlg = simpledialog.askstring("Add User", "Enter new username:")
        if not dlg: return
        pw = simpledialog.askstring("Add User", f"Enter password for {dlg}:", show='*')
        if not pw: return
        self.db.cur.execute("INSERT INTO Users(userName,Password,role) VALUES(?,?,?)",
                            dlg, hash_sha256(pw), 'user')
        messagebox.showinfo("Success", f"User {dlg} added.")

    def change_password_dialog(self):
        if self.role == 'admin':
            user_to_change = simpledialog.askstring("Change Password", "Enter username to change password:")
            if not user_to_change: return
            new_pass = simpledialog.askstring("Change Password", f"Enter new password for {user_to_change}:", show='*')
            if not new_pass: return
            self.db.cur.execute("UPDATE Users SET Password=? WHERE userName=?",
                                hash_sha256(new_pass), user_to_change)
            messagebox.showinfo("Success", f"Password for {user_to_change} updated.")
        else:
            admin_pass = simpledialog.askstring("Admin Verification", "Enter Admin password:", show='*')
            if not admin_pass: return
            self.db.cur.execute("SELECT Password FROM Users WHERE role='admin'")
            admins = self.db.cur.fetchall()
            admin_verified = any(a.Password == hash_sha256(admin_pass) for a in admins)
            if not admin_verified:
                messagebox.showerror("Permission Denied", "Invalid admin password.")
                return
            new_pass = simpledialog.askstring("Change Password", "Enter your new password:", show='*')
            if not new_pass: return
            self.db.cur.execute("UPDATE Users SET Password=? WHERE userName=?",
                                hash_sha256(new_pass), self.user)
            messagebox.showinfo("Success", "Your password has been updated.")

    def export_daily_report(self):
        date_str = simpledialog.askstring("Report Date", "Enter date (YYYY-MM-DD) or leave blank for today:")
        try:
            report_data, report_date = self.db.get_daily_user_report(date_str)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate report: {e}")
            return

        if not report_data:
            messagebox.showinfo("No Data", f"No non-admin entries on {report_date}")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"daily_report_{report_date}.xlsx"
        )
        if not filename:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Report {report_date}"
            title = f"Daily Transaction Report - {report_date}"
            ws.merge_cells('A1:I1')
            ws['A1'] = title
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')

            row = 3
            for user, entries in report_data.items():
                ws.cell(row, 1, f"User: {user}").font = Font(bold=True)
                ws.cell(row, 2, f"Total Entries: {len(entries)}").font = Font(bold=True)
                row += 1
                headers = ["ID", "Name", "Section", "Course", "Year", "Address", "Amount", "Type"]
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row, col, h)
                    cell.font = Font(bold=True)
                    cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'),
                                         left=Side(style='thin'), right=Side(style='thin'))
                    cell.alignment = Alignment(horizontal='center')
                row += 1
                for entry in entries:
                    for col, key in enumerate(["ID", "Name", "Section", "Course", "Year", "Address", "Amount", "Type"], 1):
                        val = entry[key]
                        if key == "Amount":
                            val = f"₱{val:,.2f}"
                        ws.cell(row, col, val)
                    row += 1
                row += 1

            for col in range(1, 9):
                max_len = 0
                column = get_column_letter(col)
                for cell in ws[column]:
                    try:
                        if len(str(cell.value)) > max_len:
                            max_len = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_len + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(filename)
            messagebox.showinfo("Success", f"Report saved to:\n{filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel: {e}")

    def tamper_check(self):
        while True:
            try:
                data = self.db.get_all_raw()
                for i, s in enumerate(data):
                    enc = [encrypt_data(s[k]) for k in ('name','section','course','year','address')]
                    prev = data[i-1]['hash_curr'] if i > 0 else '0'*64
                    concat = "".join(enc) + str(s['Amount']) + s['TransactionCreated'] + s['created_by'] + prev
                    expected = hash_sha256(concat)
                    if expected != s['hash_curr']:
                        msg = f"DATA TAMPERED! – transaction {s['transactionId']}"
                        self.db._notify(msg)
                        print(f"[TAMPER ALERT] {msg}")
            except Exception as e:
                print(f"[Tamper Check Error] {e}")
            time.sleep(2)

class AddTransactionDialog(tk.Toplevel):
    def __init__(self, parent, username):
        super().__init__(parent)
        self.title("Add Transaction")
        self.geometry("480x500")
        self.username = username
        self.result = False
        fields = ["Name","Section","Course","Year","Address","Amount","Transaction Type"]
        self.entries = {}
        for i, f in enumerate(fields):
            ttk.Label(self, text=f"{f}:").grid(row=i, column=0, sticky='e', padx=10, pady=5)
            if f == "Transaction Type":
                var = tk.StringVar(value="School Fee")
                combo = ttk.Combobox(self, textvariable=var, state="readonly", width=35)
                combo['values'] = [
                    "School Fee", "Miscellaneous Fee", "ID Fee", "Enrollment Fee",
                    "Uniform Fee", "COR Replacement Fee", "ID Replacement Fee",
                    "Library Fee", "Other Fees"
                ]
                combo.grid(row=i, column=1, padx=10, pady=5)
                self.entries[f] = var
            else:
                e = ttk.Entry(self, width=38)
                e.grid(row=i, column=1, padx=10, pady=5)
                self.entries[f] = e

        btns = ttk.Frame(self)
        btns.grid(row=len(fields), column=0, columnspan=2, pady=15)
        ttk.Button(btns, text="Save", command=self.save).pack(side='left', padx=6)
        ttk.Button(btns, text="Cancel", command=self.destroy).pack(side='left', padx=6)

    def save(self):
        data = {}
        for k, w in self.entries.items():
            val = w.get().strip() if hasattr(w, 'get') else w.get()
            if not val:
                messagebox.showwarning("Empty", f"{k} is required")
                return
            data[k] = val

        try:
            amount = float(data["Amount"])
        except ValueError:
            messagebox.showwarning("Invalid", "Amount must be a number")
            return

        db = Database()
        db.insert_transaction(
            data["Name"], data["Section"], data["Course"], data["Year"],
            data["Address"], amount, data["Transaction Type"], self.username
        )
        self.result = True
        self.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()